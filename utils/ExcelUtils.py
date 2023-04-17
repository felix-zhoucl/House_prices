# coding=utf-8
# @Author:zcl01
# @Create_time:2023/4/3 9:54
# @File_name:ExcelUtils
import xlsxwriter

workbook = xlsxwriter.Workbook("test.xslx")  # 创建excel文件

worksheet1 = workbook.add_worksheet('操作日志')  # 括号内为工作表表名

bold = workbook.add_format({

    'bold': True,  # 字体加粗

    'border': 1,  # 单元格边框宽度

    'align': 'left',  # 水平对齐方式

    'valign': 'vcenter',  # 垂直对齐方式

    'fg_color': '#F4B084',  # 单元格背景颜色

    'text_wrap': True,  # 是否自动换行

})

# row:行索引， col：列索引， data:要写入的数据, bold:单元格的样式
# 索引下标从0开始
worksheet1.write(1, 3, "data_1", bold)

# A1:从A1单元格开始插入数据，按行插入， data:要写入的数据（格式为一个列表), bold:单元格的样式
worksheet1.write_row("A1", "data_2", bold)

# 　A1:从A1单元格开始插入数据，按列插入， data:要写入的数据（格式为一个列表), bold:单元格的样式
worksheet1.write_column("A1", "data_3", bold)

# 　第一个参数是插入的起始单元格，第二个参数是图片你文件的绝对路径

# worksheet1.insert_image('A1', 'f:\\1.jpg')
#
# worksheet1.merge_range()  # 合并单元格

workbook.worksheets()  # 获得当前excel文件的所有工作表

workbook.close()  # 关闭excel文件


def DYNAMIC_FILL_COLOR(self, plugin_name, *args):
    """
    单元格背景颜色填充
    1深红 2红色 3橙色 4黄色 5浅绿 6绿色 7浅蓝 8蓝色 9深蓝 10紫色
    :param plugin_name: 云函数名称
    :param args: 云函数参数
    :return: {'num': 1, 'value': 12}  or {'color': '7030A0', 'value': '值'}  color 为色值
    """
    # 1深红 2红色 3橙色 4黄色 5浅绿 6绿色 7浅蓝 8蓝色 9深蓝 10紫色
    color_dict = {1: 'C00000', 2: 'FF0000', 3: 'FFC000', 4: 'FFFF00', 5: '92D050',
                  6: '00B050', 7: '00B0F0', 8: '0070C0', 9: '002060', 10: '7030A0'}
    from openpyxl.styles import PatternFill
    from openpyxl.styles import Font

    _col = self.context.get('cell_address').get('col')
    _row = self.context.get('cell_address').get('row')
    cell = '{}{}'.format(get_column_letter(_col), _row)
    plugin_res = run_dynamic_script(
        self.context['COMPANY_ID'],
        plugin_name,
        args=args or [],
        default="NO_SCRIPT"
    )

    if self.context.get('sheet_name'):
        _ind = self.context.get('workbook').workbook.sheetnames.index(self.context.get('sheet_name'))
        _ws = self.context.get('workbook').workbook.worksheets[_ind]
    else:
        _ws = self.context.get('workbook').workbook.worksheets[0]

    if plugin_res != "NO_SCRIPT":
        value_dict = plugin_res
    else:
        value_dict = {}

    num = value_dict.get('num')
    color = value_dict.get('color')
    value = value_dict.get('value')

    def set_print_style(_p_b, _c, _c_d, _s_n=None):
        if _s_n:
            _d_n = _p_b._get_div_class_name(_s_n)
        else:
            _d_n = 'common'
        key = ".{} .{} .{} .{}".format(
            _d_n, 'body', "row_{}".format(_c.get('row')), "col_{}".format(_c.get('col')))
        return _p_b.set_html_style(key, _c_d)

    font_style = value_dict.get('font_style')
    if font_style:
        font_size = font_style.get('font_size')
        bold = font_style.get('bold')
        if font_size and bold:
            if self.context.get('print_book'):
                css_dic = {
                    "font-size": "{}px".format(font_size),
                    "font-weight": "bold"
                }
                set_print_style(
                    self.context.get('print_book'),
                    self.context['cell_address'],
                    css_dic,
                    _s_n=self.context.get('sheet_name')
                )
            else:
                _style = Font(size=font_size, bold=True)
                _ws[cell].font = _style
        if bold:
            if self.context.get('print_book'):
                css_dic = {
                    "font-weight": "bold"
                }
                set_print_style(
                    self.context.get('print_book'),
                    self.context['cell_address'],
                    css_dic,
                    _s_n=self.context.get('sheet_name')
                )
            else:
                _style = Font(bold=True)
                _ws[cell].font = _style
        if font_size:
            if self.context.get('print_book'):
                css_dic = {
                    "font-size": "{}px".format(font_size)
                }
                set_print_style(
                    self.context.get('print_book'),
                    self.context['cell_address'],
                    css_dic,
                    _s_n=self.context.get('sheet_name')
                )
            else:
                _style = Font(size=font_size)
                _ws[cell].font = _style
    if color:
        if self.context.get('print_book'):
            css_dic = {
                "color": "{}".format(color),
            }
            set_print_style(
                self.context.get('print_book'),
                self.context['cell_address'],
                css_dic,
                _s_n=self.context.get('sheet_name')
            )
        else:
            fill = PatternFill("solid", fgColor=color)
            _ws[cell].fill = fill
    if num and not color:
        if self.context.get('print_book'):
            css_dic = {
                "color": "{}".format(color_dict[num]),
            }
            set_print_style(
                self.context.get('print_book'),
                self.context['cell_address'],
                css_dic,
                _s_n=self.context.get('sheet_name')
            )
        else:
            color = color_dict[num]
            fill = PatternFill("solid", fgColor=color)
            _ws[cell].fill = fill
    if value:
        return value
    else:
        return ''
